from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample4.xlsx")
ws = wb.active

# 번호 영어 수학

a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

#style 지정
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자색은 빨갛게, 이탤릭, 두껍게
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 글자색 보라, 폰트 Arial, 최소선
c1.font = Font(color="0000FF", size=20, underline="single") # 색은 파랑 글자크기 20, 밑줄적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 cell에 대해서 정렬
        cell.alignment= Alignment (horizontal="center", vertical="center")
        # center, left, right, bottom

        if cell.column == 1: # A번호열은 제외
            continue

        # cell이 정수형 데이터이고 90 점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로
            cell.font = Font(color="FF0000") # 폰트색 변경

# 틀 고정
ws.freeze_panes = "B2" # B2기준으로 틀 고정


wb.save("sample_style.xlsx")