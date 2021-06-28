from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성
# wb = active
ws = wb.create_sheet() # 새로운 sheet 기본이름으로 생성
ws.title = "My sheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경

#sheet, mysheet, yoursheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 sheet 이름 확인

#Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title =  "Copoed Sheet"

wb.save("sample2.xlsx")